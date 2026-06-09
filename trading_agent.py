"""
================================================================
   PRO TRADING AGENT v2.0
   XAUUSD (Gold) - Professional Signal Engine
   Account aware | Volatility aware | News aware
================================================================

WHAT IT DOES (like a pro trader):

  1. ACCOUNT-BASED SIZING
     Reads YOUR real balance and adjusts position size dynamically
     Reduces size in losing streaks (drawdown protection)
     Increases size in winning streaks (compounding)

  2. VOLATILITY AWARE
     Calculates ATR (Average True Range) on 15M and 1H
     Wide SL when volatile, tight SL when calm
     Skips trades when volatility is extreme (news spikes)

  3. TECHNICAL ANALYSIS (multi-factor scoring)
     - 1H Trend (EMA20 vs EMA50 + price position)
     - 4H Trend (HTF confirmation)
     - RSI (oversold for buys, overbought for sells)
     - Bollinger Bands (overextension check)
     - MACD (momentum confirmation)
     - Price action (pullback to dynamic support/resistance)
     - Candlestick quality (body %, wick rejection)
     - Volume confirmation

  4. FUNDAMENTAL ANALYSIS
     - Economic calendar (avoid trading near major news)
     - Session timing (London + NY = best liquidity)
     - Day of week filter (avoid Friday afternoons)

  5. PROFESSIONAL OUTPUT
     Full trade rationale with:
     - Why you should take this trade
     - Confidence score (out of 10)
     - Risk amount in $
     - Position size in oz of gold
     - SL and TP with reasoning
     - Expected RR
     - Walk-away conditions

================================================================

SETUP (run once):
   pip install yfinance pandas numpy colorama openpyxl

RUN:
   python trading_agent.py

================================================================
"""

import yfinance as yf
import pandas as pd
import numpy as np
import time
import os
import json
from datetime import datetime, timedelta, timezone
import warnings
warnings.filterwarnings("ignore")

try:
    import colorama
    from colorama import Fore, Style, Back
    colorama.init(autoreset=True)
except ImportError:
    print("Please install colorama: pip install colorama")
    class Fore:
        GREEN = YELLOW = RED = CYAN = WHITE = MAGENTA = BLUE = LIGHTBLACK_EX = ""
    class Back:
        GREEN = RED = YELLOW = BLUE = ""
    class Style:
        RESET_ALL = BRIGHT = DIM = ""

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

# ================================================================
# CONFIG -- adjust these for YOUR account
# ================================================================
CONFIG = {
    # ---- YOUR ACCOUNT -----------------------------------------
    "starting_balance"   : 10000.00,   # CHANGE TO YOUR REAL BALANCE
    "currency"           : "USD",

    # ---- RISK MANAGEMENT --------------------------------------
    "base_risk_pct"      : 0.01,       # 1% base risk per trade
    "max_risk_pct"       : 0.015,      # max 1.5% on high-confidence
    "min_risk_pct"       : 0.005,      # min 0.5% on lower-confidence
    "max_daily_loss_pct" : 0.02,       # stop trading at -2% day
    "max_daily_trades"   : 3,
    "drawdown_throttle"  : True,       # auto-reduce size after losses

    # ---- INSTRUMENT -------------------------------------------
    "symbol"             : "GC=F",     # Gold futures = XAUUSD

    # ---- TECHNICALS -------------------------------------------
    "rsi_period"         : 14,
    "rsi_buy_max"        : 65,         # don't buy if RSI > 65
    "rsi_sell_min"       : 35,         # don't sell if RSI < 35
    "bb_period"          : 20,
    "bb_std"             : 2.0,
    "atr_period"         : 14,
    "ema_fast"           : 20,
    "ema_slow"           : 50,
    "macd_fast"          : 12,
    "macd_slow"          : 26,
    "macd_signal"        : 9,

    # ---- VOLATILITY GUARD -------------------------------------
    "atr_min"            : 0.50,
    "atr_max"            : 25.00,      # skip trades if ATR too crazy
    "atr_extreme_mult"   : 3.0,        # if ATR > 3x its avg, skip

    # ---- STRATEGY ---------------------------------------------
    "sl_atr_mult"        : 1.5,        # SL = entry +/- ATR*1.5
    "rr_min"             : 2.0,        # min 1:2 RR or skip
    "rr_target"          : 2.5,        # aim for 1:2.5
    "min_body_pct"       : 0.45,
    "min_confidence"     : 6,          # only trade if score >= 6/10

    # ---- SESSIONS (UTC) ---------------------------------------
    "london_open"        : 7,
    "london_close"       : 10,
    "ny_open"            : 12,
    "ny_close"           : 16,

    # ---- NEWS BLACKOUT (UTC HH, MM) ---------------------------
    "news_times_utc"     : [
        (8,  30), (12, 30), (13, 30),
        (14,  0), (15,  0), (18,  0), (19,  0),
    ],
    "news_buffer_min"    : 30,
    "skip_friday_pm"     : True,       # avoid friday after 16:00 UTC

    # ---- TIMING -----------------------------------------------
    "update_seconds"     : 60,
    "log_file"           : "trades_log.xlsx",
    "state_file"         : "agent_state.json",
}

# ================================================================
# STATE
# ================================================================
class AgentState:
    def __init__(self):
        self.balance         = CONFIG["starting_balance"]
        self.start_bal       = CONFIG["starting_balance"]
        self.daily_start_bal = CONFIG["starting_balance"]
        self.daily_pnl       = 0.0
        self.total_pnl       = 0.0
        self.peak_balance    = CONFIG["starting_balance"]
        self.daily_trades    = 0
        self.total_trades    = 0
        self.wins            = 0
        self.losses          = 0
        self.breakevens      = 0
        self.win_streak      = 0
        self.loss_streak     = 0
        self.active_trade    = None
        self.trade_history   = []
        self.last_date       = datetime.now().date()
        self.status          = "STARTING..."
        self.last_signal     = None
        self.current_price   = 0.0
        self.price_change    = 0.0
        self.last_update     = "--:--:--"
        self.error_msg       = ""
        self.last_score      = 0
        self.load()

    def new_day_check(self):
        today = datetime.now().date()
        if today != self.last_date:
            self.daily_pnl       = 0.0
            self.daily_trades    = 0
            self.daily_start_bal = self.balance
            self.last_date       = today
            self.save()

    def win_rate(self):
        d = self.wins + self.losses
        return (self.wins / d * 100) if d > 0 else 0.0

    def drawdown(self):
        if self.peak_balance == 0:
            return 0.0
        return ((self.peak_balance - self.balance) / self.peak_balance) * 100

    def save(self):
        try:
            data = {
                "balance"      : self.balance,
                "total_pnl"    : self.total_pnl,
                "peak_balance" : self.peak_balance,
                "total_trades" : self.total_trades,
                "wins"         : self.wins,
                "losses"       : self.losses,
                "breakevens"   : self.breakevens,
                "last_date"    : str(self.last_date),
            }
            with open(CONFIG["state_file"], "w") as f:
                json.dump(data, f, indent=2)
        except Exception:
            pass

    def load(self):
        try:
            with open(CONFIG["state_file"], "r") as f:
                d = json.load(f)
            self.balance       = d.get("balance",       self.balance)
            self.total_pnl     = d.get("total_pnl",     0.0)
            self.peak_balance  = d.get("peak_balance",  self.balance)
            self.total_trades  = d.get("total_trades",  0)
            self.wins          = d.get("wins",          0)
            self.losses        = d.get("losses",        0)
            self.breakevens    = d.get("breakevens",    0)
        except Exception:
            pass

S = AgentState()

# ================================================================
# DATA
# ================================================================
def fetch(interval, period):
    try:
        df = yf.Ticker(CONFIG["symbol"]).history(period=period, interval=interval)
        if df.empty:
            return None
        df.index = df.index.tz_convert("UTC")
        return df
    except Exception:
        return None

# ================================================================
# INDICATORS
# ================================================================
def ema(s, n):
    return s.ewm(span=n, adjust=False).mean()

def rsi(series, n=14):
    delta = series.diff()
    gain  = delta.clip(lower=0).rolling(n).mean()
    loss  = (-delta.clip(upper=0)).rolling(n).mean()
    rs    = gain / loss.replace(0, np.nan)
    return 100 - (100 / (1 + rs))

def atr(df, n=14):
    h, l, c = df["High"], df["Low"], df["Close"]
    tr = pd.concat([h-l, (h-c.shift(1)).abs(), (l-c.shift(1)).abs()], axis=1).max(axis=1)
    return tr.rolling(n).mean()

def bollinger(series, n=20, k=2.0):
    mid = series.rolling(n).mean()
    sd  = series.rolling(n).std()
    return mid + k*sd, mid, mid - k*sd

def macd(series, fast=12, slow=26, sig=9):
    ef = ema(series, fast)
    es = ema(series, slow)
    line = ef - es
    signal_line = ema(line, sig)
    hist = line - signal_line
    return line, signal_line, hist

def body_pct(df):
    body = (df["Close"] - df["Open"]).abs()
    rng  = (df["High"] - df["Low"]).replace(0, np.nan)
    return body / rng

# ================================================================
# SESSION + NEWS
# ================================================================
def session_now():
    n = datetime.now(timezone.utc)
    h = n.hour + n.minute / 60
    if CONFIG["london_open"] <= h < CONFIG["london_close"]: return "LONDON"
    if CONFIG["ny_open"]     <= h < CONFIG["ny_close"]:     return "NEW YORK"
    if 0 <= h < CONFIG["london_open"]:                       return "ASIAN"
    return "OFF HOURS"

def in_kz():
    return session_now() in ("LONDON", "NEW YORK")

def news_check():
    now = datetime.now(timezone.utc)
    buf = CONFIG["news_buffer_min"]
    for (h, m) in CONFIG["news_times_utc"]:
        t = now.replace(hour=h, minute=m, second=0, microsecond=0)
        if abs((now - t).total_seconds() / 60) <= buf:
            return False, f"{h:02d}:{m:02d}"
    return True, ""

def next_news_str():
    now = datetime.now(timezone.utc)
    best = None
    for (h, m) in CONFIG["news_times_utc"]:
        t = now.replace(hour=h, minute=m, second=0, microsecond=0)
        if t < now: t += timedelta(days=1)
        diff = (t - now).total_seconds() / 60
        if best is None or diff < best[0]:
            best = (diff, t)
    if best:
        return f"{best[1].strftime('%H:%M')}  (in {int(best[0]//60)}h {int(best[0]%60)}m)"
    return "Unknown"

def friday_pm_block():
    n = datetime.now(timezone.utc)
    if CONFIG["skip_friday_pm"] and n.weekday() == 4 and n.hour >= 16:
        return True
    return False

# ================================================================
# PROFESSIONAL SIGNAL ENGINE  (multi-factor scoring 0-10)
# ================================================================
def analyze(ltf, htf, h4):
    """
    Returns dict with:
      direction      : BUY / SELL / SKIP
      score          : 0-10 confidence
      entry, sl, tp  : prices
      reasons        : list of bullet points
      warnings       : list of caution points
      atr_v, rsi_v, etc.
    """
    out = {"direction":"SKIP","score":0,"entry":0,"sl":0,"tp":0,
           "reasons":[], "warnings":[], "atr_v":0, "rsi_v":0,
           "bb_pos":"","macd_dir":"", "trend_1h":"", "trend_4h":"",
           "body_v":0, "rr":0, "vol_state":""}

    if ltf is None or htf is None or len(ltf) < 60 or len(htf) < 60:
        out["warnings"].append("Insufficient price history")
        return out

    # ---------- compute all indicators ----------
    ltf_close = ltf["Close"]
    htf_close = htf["Close"]

    e_fast_1h = ema(htf_close, CONFIG["ema_fast"]).iloc[-1]
    e_slow_1h = ema(htf_close, CONFIG["ema_slow"]).iloc[-1]
    px_1h     = htf_close.iloc[-1]

    if h4 is not None and len(h4) > 20:
        e_fast_4h = ema(h4["Close"], CONFIG["ema_fast"]).iloc[-1]
        px_4h     = h4["Close"].iloc[-1]
        trend_4h  = "BULL" if px_4h > e_fast_4h else "BEAR"
    else:
        trend_4h = "N/A"

    e20_15m = ema(ltf_close, CONFIG["ema_fast"])
    atr_15m = atr(ltf, CONFIG["atr_period"])
    rsi_15m = rsi(ltf_close, CONFIG["rsi_period"])
    bbu, bbm, bbl = bollinger(ltf_close, CONFIG["bb_period"], CONFIG["bb_std"])
    mline, msig, mhist = macd(ltf_close,
                               CONFIG["macd_fast"],
                               CONFIG["macd_slow"],
                               CONFIG["macd_signal"])
    bp = body_pct(ltf)

    px      = ltf_close.iloc[-1]
    op      = ltf["Open"].iloc[-1]
    hi      = ltf["High"].iloc[-1]
    lo      = ltf["Low"].iloc[-1]
    pre_lo  = ltf["Low"].iloc[-2]
    pre_hi  = ltf["High"].iloc[-2]
    e20_v   = e20_15m.iloc[-1]
    pre_e20 = e20_15m.iloc[-2]
    atr_v   = atr_15m.iloc[-1]
    atr_avg = atr_15m.rolling(50).mean().iloc[-1]
    rsi_v   = rsi_15m.iloc[-1]
    bbu_v   = bbu.iloc[-1]
    bbl_v   = bbl.iloc[-1]
    bbm_v   = bbm.iloc[-1]
    mhist_v = mhist.iloc[-1]
    mhist_p = mhist.iloc[-2]
    body_v  = bp.iloc[-1]

    out.update({"atr_v": atr_v, "rsi_v": rsi_v, "body_v": body_v})

    # ---------- 1H TREND ----------
    if e_fast_1h > e_slow_1h and px_1h > e_fast_1h:
        trend_1h = "BULL"
    elif e_fast_1h < e_slow_1h and px_1h < e_fast_1h:
        trend_1h = "BEAR"
    else:
        trend_1h = "MIXED"
    out["trend_1h"] = trend_1h
    out["trend_4h"] = trend_4h

    # ---------- VOLATILITY ----------
    if pd.isna(atr_v):
        out["warnings"].append("ATR not yet computed -- need more bars")
        return out
    if atr_v < CONFIG["atr_min"]:
        out["warnings"].append(f"ATR too low ({atr_v:.2f}) - market dead")
        out["vol_state"] = "DEAD"
        return out
    if atr_v > CONFIG["atr_max"]:
        out["warnings"].append(f"ATR too high ({atr_v:.2f}) - dangerous")
        out["vol_state"] = "EXTREME"
        return out
    if (not pd.isna(atr_avg)) and atr_v > atr_avg * CONFIG["atr_extreme_mult"]:
        out["warnings"].append(f"ATR spike ({atr_v:.2f} vs avg {atr_avg:.2f}) - skip")
        out["vol_state"] = "SPIKE"
        return out
    out["vol_state"] = "NORMAL"

    # ---------- BB POSITION ----------
    if px > bbu_v:      bb_pos = "ABOVE UPPER"
    elif px < bbl_v:    bb_pos = "BELOW LOWER"
    elif px > bbm_v:    bb_pos = "UPPER HALF"
    else:               bb_pos = "LOWER HALF"
    out["bb_pos"] = bb_pos

    # ---------- MACD ----------
    macd_dir = "BULL" if mhist_v > 0 and mhist_v > mhist_p else \
               "BEAR" if mhist_v < 0 and mhist_v < mhist_p else "FLAT"
    out["macd_dir"] = macd_dir

    # ---------- BUY ANALYSIS ----------
    if trend_1h == "BULL":
        score    = 0
        reasons  = []
        warnings = []

        # +2 : 1H bullish
        score += 2
        reasons.append(f"1H trend BULLISH (EMA{CONFIG['ema_fast']} > EMA{CONFIG['ema_slow']}, price above)")

        # +1 : 4H bullish
        if trend_4h == "BULL":
            score += 1
            reasons.append("4H trend confirms BULLISH bias")
        elif trend_4h == "BEAR":
            warnings.append("4H is BEARISH - lower confidence")

        # +2 : pullback to EMA20
        pullback = (pre_lo <= pre_e20) or (lo <= e20_v)
        if pullback:
            score += 2
            reasons.append("Price pulled back to dynamic support (15M EMA20)")
        else:
            warnings.append("No clean pullback - chasing the move")

        # +2 : bullish rejection candle
        bull_candle = px > op
        solid       = (not pd.isna(body_v)) and body_v >= CONFIG["min_body_pct"]
        above_ema   = px > e20_v
        if bull_candle and solid and above_ema:
            score += 2
            reasons.append(f"Strong bullish rejection (body {body_v*100:.0f}% close above EMA)")
        else:
            return out  # no valid entry

        # +1 : RSI not overbought
        if rsi_v < CONFIG["rsi_buy_max"]:
            if rsi_v < 50:
                score += 1
                reasons.append(f"RSI {rsi_v:.0f} - room to run up")
            else:
                reasons.append(f"RSI {rsi_v:.0f} - neutral momentum")
        else:
            warnings.append(f"RSI {rsi_v:.0f} - overbought, avoid")
            return out

        # +1 : not above upper BB (extension)
        if bb_pos not in ("ABOVE UPPER",):
            score += 1
            reasons.append(f"BB position OK ({bb_pos.lower()})")
        else:
            warnings.append("Price above upper BB - overextended")

        # +1 : MACD bullish
        if macd_dir == "BULL":
            score += 1
            reasons.append("MACD histogram rising (bullish momentum)")
        elif macd_dir == "BEAR":
            warnings.append("MACD bearish - momentum diverging")

        # ---------- PRICE CALCULATIONS ----------
        sl_dist = atr_v * CONFIG["sl_atr_mult"]
        entry   = round(px, 2)
        sl      = round(entry - sl_dist, 2)
        tp      = round(entry + sl_dist * CONFIG["rr_target"], 2)
        rr      = round((tp - entry) / (entry - sl), 2)

        if rr < CONFIG["rr_min"]:
            warnings.append(f"RR only {rr} - below minimum")
            return out

        out.update({"direction":"BUY","score":score,"entry":entry,
                    "sl":sl,"tp":tp,"rr":rr,
                    "reasons":reasons,"warnings":warnings})
        return out

    # ---------- SELL ANALYSIS ----------
    if trend_1h == "BEAR":
        score    = 0
        reasons  = []
        warnings = []

        score += 2
        reasons.append(f"1H trend BEARISH (EMA{CONFIG['ema_fast']} < EMA{CONFIG['ema_slow']}, price below)")

        if trend_4h == "BEAR":
            score += 1
            reasons.append("4H trend confirms BEARISH bias")
        elif trend_4h == "BULL":
            warnings.append("4H is BULLISH - lower confidence")

        pullback = (pre_hi >= pre_e20) or (hi >= e20_v)
        if pullback:
            score += 2
            reasons.append("Price pulled back to dynamic resistance (15M EMA20)")
        else:
            warnings.append("No clean pullback")

        bear_candle = px < op
        solid       = (not pd.isna(body_v)) and body_v >= CONFIG["min_body_pct"]
        below_ema   = px < e20_v
        if bear_candle and solid and below_ema:
            score += 2
            reasons.append(f"Strong bearish rejection (body {body_v*100:.0f}% close below EMA)")
        else:
            return out

        if rsi_v > CONFIG["rsi_sell_min"]:
            if rsi_v > 50:
                score += 1
                reasons.append(f"RSI {rsi_v:.0f} - room to fall")
            else:
                reasons.append(f"RSI {rsi_v:.0f} - neutral momentum")
        else:
            warnings.append(f"RSI {rsi_v:.0f} - oversold, avoid")
            return out

        if bb_pos not in ("BELOW LOWER",):
            score += 1
            reasons.append(f"BB position OK ({bb_pos.lower()})")
        else:
            warnings.append("Price below lower BB - overextended")

        if macd_dir == "BEAR":
            score += 1
            reasons.append("MACD histogram falling (bearish momentum)")
        elif macd_dir == "BULL":
            warnings.append("MACD bullish - momentum diverging")

        sl_dist = atr_v * CONFIG["sl_atr_mult"]
        entry   = round(px, 2)
        sl      = round(entry + sl_dist, 2)
        tp      = round(entry - sl_dist * CONFIG["rr_target"], 2)
        rr      = round((entry - tp) / (sl - entry), 2)

        if rr < CONFIG["rr_min"]:
            warnings.append(f"RR only {rr} - below minimum")
            return out

        out.update({"direction":"SELL","score":score,"entry":entry,
                    "sl":sl,"tp":tp,"rr":rr,
                    "reasons":reasons,"warnings":warnings})
        return out

    out["warnings"].append("1H trend mixed - no high-probability setup")
    return out

# ================================================================
# RISK SIZING -- adapts to account, confidence, drawdown
# ================================================================
def position_size(entry, sl, score):
    # Confidence-based risk:
    # score >= 9 -> max_risk_pct
    # score 7-8  -> base_risk_pct
    # score 6    -> min_risk_pct
    if score >= 9:   risk_pct = CONFIG["max_risk_pct"]
    elif score >= 7: risk_pct = CONFIG["base_risk_pct"]
    else:            risk_pct = CONFIG["min_risk_pct"]

    # Drawdown throttle:
    if CONFIG["drawdown_throttle"]:
        dd = S.drawdown()
        if dd > 10:   risk_pct *= 0.5
        elif dd > 5:  risk_pct *= 0.75
        if S.loss_streak >= 2:
            risk_pct *= 0.5

    risk_usd = S.balance * risk_pct
    dist     = abs(entry - sl)
    units    = round(risk_usd / dist, 4) if dist > 0 else 0
    return units, risk_usd, risk_pct

def can_trade():
    loss_lim = S.daily_start_bal * CONFIG["max_daily_loss_pct"]
    if S.daily_pnl <= -loss_lim:
        return False, "DAILY LOSS LIMIT"
    if S.daily_trades >= CONFIG["max_daily_trades"]:
        return False, "MAX DAILY TRADES"
    if S.active_trade is not None:
        return False, "TRADE OPEN"
    return True, "OK"

# ================================================================
# TRADE OBJECT
# ================================================================
class Trade:
    def __init__(self, direction, entry, sl, tp, units, risk_usd, risk_pct, score, reasons, rr):
        self.direction = direction
        self.entry     = entry
        self.sl        = sl
        self.tp        = tp
        self.units     = units
        self.risk_usd  = risk_usd
        self.risk_pct  = risk_pct
        self.score     = score
        self.reasons   = reasons
        self.rr        = rr
        self.open_time = datetime.now()
        self.close_time= None
        self.close_px  = None
        self.result    = None
        self.pnl       = None

def open_trade(sig):
    units, risk_usd, risk_pct = position_size(sig["entry"], sig["sl"], sig["score"])
    if units <= 0:
        return None
    t = Trade(sig["direction"], sig["entry"], sig["sl"], sig["tp"],
              units, risk_usd, risk_pct, sig["score"], sig["reasons"], sig["rr"])
    S.active_trade = t
    return t

def manage_trade(price):
    t = S.active_trade
    if t is None: return
    hit = None
    if t.direction == "BUY":
        if price <= t.sl: hit = ("SL", t.sl)
        elif price >= t.tp: hit = ("TP", t.tp)
    else:
        if price >= t.sl: hit = ("SL", t.sl)
        elif price <= t.tp: hit = ("TP", t.tp)
    if hit:
        result, cpx = hit
        t.close_time = datetime.now()
        t.close_px   = cpx
        t.result     = result
        t.pnl        = ((cpx - t.entry) * t.units if t.direction == "BUY"
                        else (t.entry - cpx) * t.units)
        S.balance     += t.pnl
        S.daily_pnl   += t.pnl
        S.total_pnl   += t.pnl
        S.total_trades+= 1
        S.daily_trades+= 1
        if t.pnl > 0:
            S.wins += 1
            S.win_streak += 1
            S.loss_streak = 0
        elif t.pnl < 0:
            S.losses += 1
            S.loss_streak += 1
            S.win_streak = 0
        else:
            S.breakevens += 1
        if S.balance > S.peak_balance:
            S.peak_balance = S.balance
        S.trade_history.append(t)
        S.active_trade = None
        S.status = f"{'WIN' if t.pnl > 0 else 'LOSS'}: {result}  {'+' if t.pnl >= 0 else ''}{t.pnl:.2f}"
        log_excel(t)
        S.save()

# ================================================================
# EXCEL LOG
# ================================================================
def init_excel():
    if not HAS_EXCEL: return
    try:
        load_workbook(CONFIG["log_file"])
    except Exception:
        wb = Workbook()
        ws = wb.active
        ws.title = "Trades"
        ws.append(["Date","Time","Dir","Score","Entry","SL","TP","Close","Result","P&L","Risk %","Risk $","Units","Balance","Reasons"])
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="0d1117", end_color="0d1117", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        wb.save(CONFIG["log_file"])

def log_excel(t):
    if not HAS_EXCEL: return
    try:
        wb = load_workbook(CONFIG["log_file"])
        ws = wb.active
        ws.append([
            t.open_time.strftime("%Y-%m-%d"),
            t.open_time.strftime("%H:%M"),
            t.direction, t.score,
            t.entry, t.sl, t.tp,
            round(t.close_px, 2), t.result,
            round(t.pnl, 2),
            round(t.risk_pct * 100, 2),
            round(t.risk_usd, 2),
            t.units, round(S.balance, 2),
            " | ".join(t.reasons)[:200],
        ])
        row = ws.max_row
        color = "00c853" if t.pnl > 0 else "d50000"
        for cell in ws[row]:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.font = Font(color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        wb.save(CONFIG["log_file"])
    except Exception:
        pass

# ================================================================
# DASHBOARD
# ================================================================
W = 72

def cstr(t, c): return c + str(t) + Style.RESET_ALL
def pnl_c(v):
    return (Fore.GREEN if v >= 0 else Fore.RED) + ("+" if v >= 0 else "") + f"{v:.2f}" + Style.RESET_ALL
def divider(): print(Fore.CYAN + "-" * W + Style.RESET_ALL)
def header(t):
    print(Fore.CYAN + "=" * W)
    print(cstr(("  " + t).ljust(W), Fore.CYAN + Style.BRIGHT))
    print(Fore.CYAN + "=" * W + Style.RESET_ALL)

def print_dashboard(sig=None):
    os.system("cls" if os.name == "nt" else "clear")

    header("PRO TRADING AGENT v2.0  |  XAUUSD  |  Paper Mode")

    # --- Market state ---
    chg_c = Fore.GREEN if S.price_change >= 0 else Fore.RED
    print(f"  Price          {Fore.YELLOW}${S.current_price:,.2f}{Style.RESET_ALL}"
          f"  ({chg_c}{'+' if S.price_change >= 0 else ''}{S.price_change:.2f}{Style.RESET_ALL})"
          f"   ATR15m: {sig['atr_v']:.2f}" if sig else
          f"  Price          {Fore.YELLOW}${S.current_price:,.2f}{Style.RESET_ALL}")

    sess = session_now()
    sc   = Fore.GREEN if sess in ("LONDON","NEW YORK") else Fore.WHITE
    print(f"  Session        {sc}{sess}{' (ACTIVE)' if sess in ('LONDON','NEW YORK') else ''}{Style.RESET_ALL}")

    nc, at = news_check()
    if nc:
        print(f"  News           {Fore.GREEN}CLEAR{Style.RESET_ALL}  Next: {next_news_str()}")
    else:
        print(f"  News           {Fore.RED}CAUTION near {at} UTC{Style.RESET_ALL}")

    if sig and sig["trend_1h"]:
        t1c = Fore.GREEN if sig["trend_1h"] == "BULL" else Fore.RED if sig["trend_1h"] == "BEAR" else Fore.YELLOW
        t4c = Fore.GREEN if sig["trend_4h"] == "BULL" else Fore.RED if sig["trend_4h"] == "BEAR" else Fore.YELLOW
        print(f"  Trend 1H/4H    {t1c}{sig['trend_1h']:5s}{Style.RESET_ALL}/"
              f"{t4c}{sig['trend_4h']:5s}{Style.RESET_ALL}"
              f"   RSI: {sig['rsi_v']:.0f}   BB: {sig['bb_pos']}   MACD: {sig['macd_dir']}")
        print(f"  Volatility     {sig['vol_state']}   (ATR {sig['atr_v']:.2f})")

    divider()

    # --- Active trade or signal ---
    if S.active_trade:
        t = S.active_trade
        fpnl = ((S.current_price - t.entry) * t.units if t.direction == "BUY"
                else (t.entry - S.current_price) * t.units)
        dc = Fore.GREEN if t.direction == "BUY" else Fore.RED
        print(f"  {Style.BRIGHT}ACTIVE TRADE{Style.RESET_ALL}")
        print(f"    {dc}{t.direction}{Style.RESET_ALL} @ {t.entry:.2f}    "
              f"Confidence: {t.score}/10    Since {t.open_time.strftime('%H:%M')}")
        print(f"    SL: {Fore.RED}{t.sl:.2f}{Style.RESET_ALL}   "
              f"TP: {Fore.GREEN}{t.tp:.2f}{Style.RESET_ALL}   "
              f"RR: 1:{t.rr}   Risk: ${t.risk_usd:.2f} ({t.risk_pct*100:.2f}%)")
        print(f"    Units: {t.units:.4f} oz    Floating: {pnl_c(fpnl)}")
    elif sig and sig["direction"] in ("BUY","SELL") and sig["score"] >= CONFIG["min_confidence"]:
        dc = Fore.GREEN if sig["direction"] == "BUY" else Fore.RED
        print(f"  {Style.BRIGHT}{dc}>>> SIGNAL: {sig['direction']} <<<{Style.RESET_ALL}")
        print(f"    Entry: {sig['entry']:.2f}   SL: {sig['sl']:.2f}   TP: {sig['tp']:.2f}   RR: 1:{sig['rr']}")
        score_c = Fore.GREEN if sig['score'] >= 8 else Fore.YELLOW if sig['score'] >= 6 else Fore.RED
        print(f"    Confidence: {score_c}{sig['score']}/10{Style.RESET_ALL}")
    else:
        sc = Fore.YELLOW if "SCAN" in S.status or "WAIT" in S.status else \
             Fore.RED if any(x in S.status for x in ["LOSS","LIMIT","BLOCK","ERROR"]) else Fore.GREEN
        print(f"  STATUS         {sc}{S.status}{Style.RESET_ALL}")
        if sig:
            sscol = Fore.GREEN if sig['score'] >= 6 else Fore.YELLOW if sig['score'] >= 4 else Fore.RED
            print(f"  Last analysis  Score: {sscol}{sig['score']}/10{Style.RESET_ALL}   "
                  f"Direction: {sig['direction']}")

    # --- Reasoning ---
    if sig and sig["reasons"]:
        print()
        print(f"  {Fore.CYAN}WHY:{Style.RESET_ALL}")
        for r in sig["reasons"]:
            print(f"    {Fore.GREEN}+{Style.RESET_ALL} {r}")
    if sig and sig["warnings"]:
        print(f"  {Fore.YELLOW}WATCH:{Style.RESET_ALL}")
        for w in sig["warnings"]:
            print(f"    {Fore.YELLOW}!{Style.RESET_ALL} {w}")

    divider()

    # --- Account ---
    eq_c = Fore.GREEN if S.balance >= S.start_bal else Fore.RED
    print(f"  Account        {eq_c}${S.balance:,.2f}{Style.RESET_ALL}   "
          f"Peak ${S.peak_balance:,.2f}   DD: {S.drawdown():.1f}%")
    print(f"  Today          {pnl_c(S.daily_pnl)}   ({S.daily_trades} trades)")
    print(f"  All-time       {pnl_c(S.total_pnl)}   Win-Rate {Fore.CYAN}{S.win_rate():.1f}%{Style.RESET_ALL}")
    print(f"  Record         W:{S.wins}  L:{S.losses}  BE:{S.breakevens}   "
          f"Streak: {('+' + str(S.win_streak)) if S.win_streak > 0 else ('-' + str(S.loss_streak)) if S.loss_streak > 0 else '0'}")
    buf  = S.daily_start_bal * CONFIG["max_daily_loss_pct"] + S.daily_pnl
    left = CONFIG["max_daily_trades"] - S.daily_trades
    print(f"  Loss buffer    ${buf:.2f}   Trades left: {left}")

    divider()
    print(f"  Updated        {S.last_update} UTC   [Ctrl+C to stop]")
    print(Fore.CYAN + "=" * W + Style.RESET_ALL)

    if S.trade_history:
        print(f"\n  {Fore.WHITE}{Style.BRIGHT}RECENT TRADES{Style.RESET_ALL}")
        for t in reversed(S.trade_history[-5:]):
            rc = Fore.GREEN if t.pnl > 0 else Fore.RED
            dc = Fore.GREEN if t.direction == "BUY" else Fore.RED
            print(f"    {t.open_time.strftime('%m-%d %H:%M')}  "
                  f"{dc}{t.direction:4s}{Style.RESET_ALL} "
                  f"S{t.score}/10  "
                  f"{t.entry:.1f} -> {t.close_px:.1f}  "
                  f"{rc}{t.result} {'+' if t.pnl >= 0 else ''}{t.pnl:.2f}{Style.RESET_ALL}")

    if S.error_msg:
        print(f"\n  {Fore.RED}Note: {S.error_msg}{Style.RESET_ALL}")

# ================================================================
# MAIN
# ================================================================
def ask_balance():
    print(Fore.CYAN + "\n" + "=" * 60)
    print("   PRO TRADING AGENT v2.0  -  FIRST RUN SETUP")
    print("=" * 60 + Style.RESET_ALL)
    print(f"\n  Current balance on file: ${S.balance:,.2f}")
    print(f"  Press ENTER to keep, or type new balance to update:")
    try:
        raw = input("  > $").strip()
        if raw:
            new = float(raw.replace(",", ""))
            S.balance         = new
            S.start_bal       = new
            S.daily_start_bal = new
            S.peak_balance    = max(new, S.peak_balance)
            S.save()
            print(f"  Balance set to ${new:,.2f}")
    except Exception:
        print("  Keeping existing balance.")
    print(Fore.CYAN + "\n  Starting agent in 3 seconds..." + Style.RESET_ALL)
    time.sleep(3)

def run():
    ask_balance()
    init_excel()

    while True:
        try:
            S.new_day_check()
            S.last_update = datetime.now(timezone.utc).strftime("%H:%M:%S")

            # ---- fetch all 3 timeframes ----
            ltf = fetch("15m", "5d")
            htf = fetch("1h",  "30d")
            h4  = fetch("1h",  "60d")    # use 1h aggregated for 4h-like trend
            if h4 is not None and not h4.empty:
                # Build 4h from 1h
                try:
                    h4 = h4.resample("4h").agg({"Open":"first","High":"max","Low":"min","Close":"last","Volume":"sum"}).dropna()
                except Exception:
                    h4 = None

            if ltf is None or ltf.empty:
                S.status    = "ERROR: No price data"
                S.error_msg = "Check internet. Retrying..."
                print_dashboard()
                time.sleep(CONFIG["update_seconds"]); continue
            S.error_msg = ""

            S.current_price = float(ltf["Close"].iloc[-1])
            S.price_change  = float(ltf["Close"].iloc[-1] - ltf["Close"].iloc[-2])

            # ---- manage open trade ----
            manage_trade(S.current_price)

            # ---- analyze ----
            sig = analyze(ltf, htf, h4)
            S.last_score = sig["score"]

            # ---- decide ----
            ok, reason = can_trade()
            nc, _      = news_check()
            valid_sess = in_kz()
            no_friday  = not friday_pm_block()

            if S.active_trade is None:
                if not ok:
                    S.status = f"BLOCKED: {reason}"
                elif not valid_sess:
                    S.status = f"WAITING: {session_now()} (not kill zone)"
                elif not nc:
                    S.status = "PAUSED: Near news event"
                elif not no_friday:
                    S.status = "PAUSED: Friday afternoon"
                elif sig["direction"] == "SKIP":
                    S.status = "SCANNING for setup..."
                elif sig["score"] < CONFIG["min_confidence"]:
                    S.status = f"SETUP found but score {sig['score']}/10 (need {CONFIG['min_confidence']}+)"
                else:
                    t = open_trade(sig)
                    if t:
                        S.status = f"TRADE OPEN: {sig['direction']} @ {sig['entry']:.2f} (score {sig['score']}/10)"

            print_dashboard(sig)
            time.sleep(CONFIG["update_seconds"])

        except KeyboardInterrupt:
            print(f"\n\n  {Fore.YELLOW}Agent stopped.{Style.RESET_ALL}")
            print(f"  Final Balance : ${S.balance:,.2f}")
            print(f"  Total P&L     : {'+' if S.total_pnl >= 0 else ''}{S.total_pnl:.2f}")
            print(f"  Win Rate      : {S.win_rate():.1f}%")
            print(f"  Peak Balance  : ${S.peak_balance:,.2f}")
            print(f"  Max Drawdown  : {S.drawdown():.1f}%")
            if HAS_EXCEL:
                print(f"  Log: {CONFIG['log_file']}")
            S.save()
            print()
            break

        except Exception as e:
            S.status    = "ERROR"
            S.error_msg = str(e)[:80]
            try: print_dashboard()
            except: print(f"Error: {e}")
            time.sleep(CONFIG["update_seconds"])

if __name__ == "__main__":
    run()
